#!/usr/bin/env python3
"""
CretePulse — HCAA (ypa.gr) Crete airports monthly traffic article generator.

Downloads the annual "ΚΙΝΗΣΗ ΑΕΡΟΛΙΜΕΝΩΝ" XLSX from HCAA (which contains
12 monthly sheets that get populated through the year, ~2 months delay),
extracts pax totals for the 3 Crete airports HER, CHQ, JSH, picks the
latest populated month, generates a bilingual EN+FR article and inserts
as draft.

This complements `fraport-chq-traffic.py` by adding HER + JSH coverage,
giving full Crete air-traffic picture.

Usage:
  python3 hcaa-crete-airports.py
  python3 hcaa-crete-airports.py --dry-run
  python3 hcaa-crete-airports.py --force
"""

import argparse
import calendar
import hashlib
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
from datetime import datetime, timezone
from urllib.parse import quote

import openpyxl
import psycopg2
import requests
from dotenv import load_dotenv

load_dotenv("/opt/cretepulse-db/.env")
load_dotenv("/opt/cretepulse/.env")

LOG_DIR = "/var/log/cretepulse-content"
LOG_FILE = f"{LOG_DIR}/hcaa-crete.log"
STOP_FILE = "/opt/cretepulse-content/.STOP"
CLAUDE_WRAPPER = "/opt/cretepulse-content/bin/claude-capped.sh"
CACHE_DIR = "/opt/cretepulse-data/hcaa"

os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

DB_PARAMS = {
    "host": "localhost",
    "port": 5433,
    "dbname": "cretepulse",
    "user": "postgres",
    "password": os.environ.get("POSTGRES_PASSWORD", ""),
}

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/121.0 Safari/537.36"
)

# Column indexes inside HCAA monthly sheet (0-based, validated 2025-12 sheet):
COL_IATA = 1
COL_NAME = 2
COL_PAX_DISEMBARK_TOTAL = 35   # Αποβιβάσεις Συνολικές
COL_PAX_EMBARK_TOTAL = 36      # Επιβιβάσεις Συνολικές
COL_PAX_TRANSIT_TOTAL = 37     # Διελεύσεις Συνολικές
COL_PAX_GRAND_TOTAL = 38       # Σύνολο

# Aircraft movements columns (0-based):
COL_AC_INTL_TOTAL = 5          # Συνολικές Εξωτερικού
COL_AC_DOM_TOTAL = 6           # Εσωτερικού Συνολικές
COL_AC_GRAND_TOTAL = 9         # Αερογραμμές

CRETE_AIRPORTS = {
    "HER": {"name_en": "Heraklion International (HER)", "name_fr": "Héraklion (HER)"},
    "CHQ": {"name_en": "Chania International (CHQ)", "name_fr": "La Canée (CHQ)"},
    "JSH": {"name_en": "Sitia (JSH)", "name_fr": "Sitia (JSH)"},
}

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

MONTH_NAMES_FR = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril",
    5: "mai", 6: "juin", 7: "juillet", 8: "août",
    9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre",
}


# --------------------------------------------------------------------------
# Logging
# --------------------------------------------------------------------------

def log(msg: str) -> None:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    line = f"{ts} [hcaa-crete] {msg}"
    print(line, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except OSError:
        pass


def check_stop() -> None:
    if os.path.exists(STOP_FILE):
        log("STOP file present — exiting.")
        sys.exit(98)


def send_telegram(text: str) -> None:
    try:
        sys.path.insert(0, "/opt/cretepulse")
        from kairos_telegram import send, Bot  # type: ignore
        send(Bot.PLUME, "HCAA Crete Airports", text)
    except Exception as e:
        log(f"Telegram error: {e}")


# --------------------------------------------------------------------------
# DB schema
# --------------------------------------------------------------------------

DDL = """
CREATE TABLE IF NOT EXISTS hcaa_crete_monthly (
    year INT NOT NULL,
    month INT NOT NULL,
    airport_iata TEXT NOT NULL,
    pax_disembark INT,
    pax_embark INT,
    pax_transit INT,
    pax_total INT,
    aircraft_intl INT,
    aircraft_dom INT,
    aircraft_total INT,
    source_xlsx_url TEXT,
    ingested_at TIMESTAMPTZ DEFAULT now(),
    PRIMARY KEY (year, month, airport_iata)
);

CREATE TABLE IF NOT EXISTS hcaa_articles_ledger (
    year INT NOT NULL,
    month INT NOT NULL,
    guide_id BIGINT,
    slug TEXT,
    generated_at TIMESTAMPTZ DEFAULT now(),
    PRIMARY KEY (year, month)
);
"""


def ensure_schema(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(DDL)
    conn.commit()


# --------------------------------------------------------------------------
# Discovery
# --------------------------------------------------------------------------

def _browser_headers(referer: str = "https://www.ypa.gr/") -> dict:
    return {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9,el;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": referer,
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
    }


def discover_xlsx_url(year: int) -> str:
    """Scrape ypa.gr stats year index for the 'KINISI' XLSX URL."""
    url = f"https://www.ypa.gr/profile/statistics/{year}/"
    r = requests.get(url, headers=_browser_headers(), timeout=30,
                     allow_redirects=True)
    r.raise_for_status()
    html = r.text
    # Pattern: link containing '%ce%9a%ce%99%ce%9d%ce%97%ce%a3%ce%97' (ΚΙΝΗΣΗ) and the year
    matches = re.findall(r'href="([^"]+\.xlsx)"', html, flags=re.I)
    candidates = [m for m in matches if "ce%9a%ce%99%ce%9d%ce%97%ce%a3%ce%97" in m.lower()
                  and str(year) in m]
    if not candidates:
        # Fallback: any KINISI/ΚΙΝΗΣΗ XLSX with the year
        candidates = [m for m in matches
                      if str(year) in m and (".xlsx" in m.lower())
                      and ("kinisi" in m.lower() or "%ce%9a" in m.lower())]
    if not candidates:
        raise RuntimeError(f"No KINISI XLSX found on {url}")
    return candidates[0]


# --------------------------------------------------------------------------
# Download + parse
# --------------------------------------------------------------------------

def download_xlsx(url: str, dest_path: str) -> None:
    if os.path.exists(dest_path) and os.path.getsize(dest_path) > 1024:
        return
    r = requests.get(url, headers=_browser_headers(), timeout=60)
    r.raise_for_status()
    tmp = dest_path + ".tmp"
    with open(tmp, "wb") as f:
        f.write(r.content)
    os.replace(tmp, dest_path)
    log(f"  saved {dest_path} ({len(r.content)} bytes)")


def to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(float(s.replace(",", "").replace(" ", "")))
    except ValueError:
        return 0


def parse_month_sheet(wb, month_num: int) -> dict:
    """Extract HER/CHQ/JSH rows from a monthly sheet. Returns
       {iata: {pax_total, pax_disembark, pax_embark, pax_transit,
               aircraft_intl, aircraft_dom, aircraft_total}}.
       If a sheet has no real data, returns {}.
    """
    sheet_name = f"{month_num:02d}"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    found = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= COL_PAX_GRAND_TOTAL:
            continue
        iata = str(row[COL_IATA]).strip() if row[COL_IATA] else ""
        if iata not in CRETE_AIRPORTS:
            continue
        pax_total = to_int(row[COL_PAX_GRAND_TOTAL])
        # Skip if all data is zero (sheet not filled yet for that month)
        found[iata] = {
            "pax_disembark": to_int(row[COL_PAX_DISEMBARK_TOTAL]),
            "pax_embark": to_int(row[COL_PAX_EMBARK_TOTAL]),
            "pax_transit": to_int(row[COL_PAX_TRANSIT_TOTAL]),
            "pax_total": pax_total,
            "aircraft_intl": to_int(row[COL_AC_INTL_TOTAL]),
            "aircraft_dom": to_int(row[COL_AC_DOM_TOTAL]),
            "aircraft_total": to_int(row[COL_AC_GRAND_TOTAL]),
        }

    # Detect if sheet is unfilled: all 3 airports have pax_total=0
    if found and all(v["pax_total"] == 0 for v in found.values()):
        return {}
    return found


def find_latest_month_with_data(wb) -> int:
    """Iterate from December back to January, return first month_num
       with non-zero data for at least HER (the busiest)."""
    for m in range(12, 0, -1):
        data = parse_month_sheet(wb, m)
        if data and data.get("HER", {}).get("pax_total", 0) > 0:
            return m
    return 0


# --------------------------------------------------------------------------
# DB upsert
# --------------------------------------------------------------------------

def upsert_airport_row(conn, year: int, month: int, iata: str, vals: dict, url: str) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO hcaa_crete_monthly (
                year, month, airport_iata,
                pax_disembark, pax_embark, pax_transit, pax_total,
                aircraft_intl, aircraft_dom, aircraft_total,
                source_xlsx_url
            )
            VALUES (%s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s)
            ON CONFLICT (year, month, airport_iata) DO UPDATE SET
                pax_disembark = EXCLUDED.pax_disembark,
                pax_embark = EXCLUDED.pax_embark,
                pax_transit = EXCLUDED.pax_transit,
                pax_total = EXCLUDED.pax_total,
                aircraft_intl = EXCLUDED.aircraft_intl,
                aircraft_dom = EXCLUDED.aircraft_dom,
                aircraft_total = EXCLUDED.aircraft_total,
                source_xlsx_url = EXCLUDED.source_xlsx_url,
                ingested_at = now()
        """, (
            year, month, iata,
            vals["pax_disembark"], vals["pax_embark"], vals["pax_transit"], vals["pax_total"],
            vals["aircraft_intl"], vals["aircraft_dom"], vals["aircraft_total"],
            url,
        ))
    conn.commit()


def fetch_year_history(conn, year: int) -> dict:
    """Return monthly totals for this year so far per airport."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT month, airport_iata, pax_total
            FROM hcaa_crete_monthly
            WHERE year=%s
            ORDER BY month, airport_iata
        """, (year,))
        rows = cur.fetchall()
    out = {}
    for m, iata, total in rows:
        out.setdefault(m, {})[iata] = total
    return out


def already_in_ledger(conn, year: int, month: int) -> bool:
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM hcaa_articles_ledger WHERE year=%s AND month=%s",
                    (year, month))
        return cur.fetchone() is not None


def record_ledger(conn, year: int, month: int, guide_id: int, slug: str) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO hcaa_articles_ledger (year, month, guide_id, slug)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (year, month) DO UPDATE SET
                guide_id = EXCLUDED.guide_id,
                slug = EXCLUDED.slug,
                generated_at = now()
        """, (year, month, guide_id, slug))
    conn.commit()


# --------------------------------------------------------------------------
# Article generation
# --------------------------------------------------------------------------

GENERATION_SYSTEM = """You are a senior bilingual travel + aviation journalist
for crete.direct, a guide to Crete, Greece. You will receive official monthly
traffic figures from the Hellenic Civil Aviation Authority (HCAA / ypa.gr)
covering the THREE Crete airports: Heraklion (HER), Chania (CHQ), and
Sitia (JSH), plus the running 12-month context for the year.

Your job: write a short, accurate, SEO-friendly article in BOTH English AND
French explaining the latest month's air traffic picture for Crete as a
whole — total passengers, split by airport, share of international flights.

CRITICAL rules:
- Base every numeric claim on the provided dataset. NEVER invent figures.
- Cite HCAA (ypa.gr) as the source AND the month covered.
- 450-650 words for body content (each language).
- No em dashes (—) and no en dashes (–). Use comma, period, " - ".
- French content: vouvoiement (vous), full accents (é è à ç ê î ô û ù).
- Tone: clear, journalistic, traveler-friendly. Light implications about
  Crete tourism momentum only if grounded in the data.
- Heraklion is the dominant airport (typically 60-70% of Crete pax),
  Chania second, Sitia very small (<1%).
- HTML body: <p>, <h2>, <ul><li>. No tables.

Output STRICT JSON ONLY (no markdown fences, no preamble):

{
  "title_en": "string, 60-95 chars, SEO headline English",
  "title_fr": "string, 60-100 chars, French headline with accents",
  "meta_desc_en": "string, 140-165 chars, English meta description",
  "meta_desc_fr": "string, 140-170 chars, French meta description",
  "slug_hint": "string, 3-6 words, lowercase hyphenated, English",
  "html_body_en": "string, valid HTML <p><h2><ul><li>. 450-650 words English",
  "html_body_fr": "string, valid HTML <p><h2><ul><li>. 450-650 words French",
  "faqs_en": [{"q":"...","a":"..."}, {"q":"...","a":"..."}, {"q":"...","a":"..."}],
  "faqs_fr": [{"q":"...","a":"..."}, {"q":"...","a":"..."}, {"q":"...","a":"..."}],
  "category": "tourism",
  "confidence": "high | medium | low"
}
"""


def build_prompt(year: int, month: int, latest_data: dict, history: dict, url: str) -> str:
    crete_total = sum(v["pax_total"] for v in latest_data.values())
    payload = {
        "source": "Hellenic Civil Aviation Authority (HCAA / ypa.gr) — ΚΙΝΗΣΗ ΑΕΡΟΛΙΜΕΝΩΝ",
        "source_xlsx_url": url,
        "month": MONTH_NAMES[month],
        "month_fr": MONTH_NAMES_FR[month],
        "year": year,
        "crete_total_pax_latest_month": crete_total,
        "airports_latest_month": latest_data,
        "year_to_date_history_by_month": history,
    }
    return (
        f"{GENERATION_SYSTEM}\n\n"
        f"---\nDATA:\n```json\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}\n```\n\n"
        f"Return strict JSON now."
    )


def call_claude(prompt: str, model: str = "haiku", timeout: int = 200) -> str:
    cmd = [
        CLAUDE_WRAPPER, "-p", prompt,
        "--model", model,
        "--output-format", "json",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if result.returncode == 98:
        raise RuntimeError("STOP file present.")
    if result.returncode == 99:
        raise RuntimeError("Claude daily cap reached.")
    if result.returncode != 0:
        raise RuntimeError(
            f"claude wrapper rc={result.returncode}: {result.stderr[:300]}"
        )
    raw = result.stdout
    try:
        wrapper = json.loads(raw)
        return (wrapper.get("result") or raw).strip()
    except json.JSONDecodeError:
        return raw.strip()


def parse_json_response(raw: str) -> dict:
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    raw = re.sub(r"\s*```$", "", raw.strip())
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            raise
        return json.loads(m.group())


# --------------------------------------------------------------------------
# Insert draft
# --------------------------------------------------------------------------

def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:90] or "crete-airports-traffic"


def derive_slug(article: dict, year: int, month: int) -> str:
    hint = article.get("slug_hint") or article.get("title_en") or "crete-airports-traffic"
    base = slugify(hint)
    h = hashlib.sha1(f"hcaa-{year}-{month}".encode()).hexdigest()[:6]
    return f"{base}-{year}-{month:02d}-{h}"


def insert_draft(conn, article: dict, year: int, month: int, url: str) -> int:
    slug = derive_slug(article, year, month)
    title_en = (article.get("title_en") or "").strip()
    title_fr = (article.get("title_fr") or "").strip()
    meta_en = (article.get("meta_desc_en") or "").strip()
    meta_fr = (article.get("meta_desc_fr") or "").strip()
    body_en = article.get("html_body_en") or ""
    body_fr = article.get("html_body_fr") or ""
    faqs_en = article.get("faqs_en") or []
    faqs_fr = article.get("faqs_fr") or []
    category = article.get("category") or "tourism"

    if not title_fr or not body_fr:
        raise RuntimeError("Missing FR fields — translator-batch needs EN+FR.")

    keywords = ["crete-airports", "hcaa", "ypa", f"{year}-{month:02d}"][:8]

    titles = {"en": title_en, "fr": title_fr}
    meta_descs = {"en": meta_en, "fr": meta_fr}
    contents = {"en": body_en, "fr": body_fr}
    faqs_obj = {"en": faqs_en, "fr": faqs_fr}

    attr_en = (
        f'<p class="source-attribution"><em>Source: '
        f'<a href="https://www.ypa.gr/profile/statistics/{year}/" rel="noopener" '
        f'target="_blank">Hellenic Civil Aviation Authority (HCAA)</a> '
        f'monthly airports movement, {MONTH_NAMES[month]} {year}.</em></p>'
    )
    attr_fr = (
        f'<p class="source-attribution"><em>Source : '
        f'<a href="https://www.ypa.gr/profile/statistics/{year}/" rel="noopener" '
        f'target="_blank">Autorité hellénique de l\'aviation civile (HCAA)</a>, '
        f"mouvement mensuel des aéroports, {MONTH_NAMES_FR[month]} {year}.</em></p>"
    )
    if attr_en not in contents["en"]:
        contents["en"] = contents["en"].rstrip() + "\n" + attr_en
    if attr_fr not in contents["fr"]:
        contents["fr"] = contents["fr"].rstrip() + "\n" + attr_fr

    word_count = len(re.sub(r"<[^>]+>", " ", contents["en"]).split())
    read_time = max(2, round(word_count / 220))

    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO guides (slug, format, category, keywords, titles, meta_descs,
                               contents, faqs, image_url, read_time, status,
                               published_at, created_at, updated_at)
            VALUES (%s, 'long', %s, %s::text[], %s::jsonb, %s::jsonb,
                    %s::jsonb, %s::jsonb, %s, %s, 'draft', now(), now(), now())
            ON CONFLICT (slug) DO UPDATE SET
                titles = EXCLUDED.titles,
                meta_descs = EXCLUDED.meta_descs,
                contents = EXCLUDED.contents,
                faqs = EXCLUDED.faqs,
                status = 'draft',
                updated_at = now()
            RETURNING id
        """, (
            slug, category, keywords,
            json.dumps(titles, ensure_ascii=False),
            json.dumps(meta_descs, ensure_ascii=False),
            json.dumps(contents, ensure_ascii=False),
            json.dumps(faqs_obj, ensure_ascii=False),
            None, read_time,
        ))
        guide_id = cur.fetchone()[0]
    conn.commit()
    return guide_id


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--year", type=int, help="Override year (default: latest)")
    args = parser.parse_args()

    check_stop()
    log(f"Run start dry_run={args.dry_run} force={args.force}")

    target_year = args.year or datetime.now(timezone.utc).year

    # Try current year first; fallback to previous if current XLSX not yet up.
    last_err = None
    for year_try in [target_year, target_year - 1]:
        try:
            url = discover_xlsx_url(year_try)
            log(f"  XLSX found for year {year_try}: {url}")
            break
        except RuntimeError as e:
            last_err = e
            log(f"  no XLSX for {year_try}: {e}")
            url = None
    if url is None:
        raise last_err or RuntimeError("No XLSX discovered")

    year = year_try
    xlsx_path = os.path.join(CACHE_DIR, f"kinisi_{year}.xlsx")
    download_xlsx(url, xlsx_path)

    log("Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    latest_month = find_latest_month_with_data(wb)
    log(f"  latest populated month: {latest_month or 'NONE'}")
    if latest_month == 0:
        log("  no populated month found; trying previous year if not done")
        if year == target_year:
            log("  retrying with previous year")
            url2 = discover_xlsx_url(year - 1)
            xlsx_path2 = os.path.join(CACHE_DIR, f"kinisi_{year-1}.xlsx")
            download_xlsx(url2, xlsx_path2)
            wb = openpyxl.load_workbook(xlsx_path2, read_only=True, data_only=True)
            latest_month = find_latest_month_with_data(wb)
            year = year - 1
            url = url2
            log(f"  fallback year {year} latest month: {latest_month}")
        if latest_month == 0:
            log("Still no data — exiting.")
            return 1

    latest_data = parse_month_sheet(wb, latest_month)
    log(f"  data: {latest_data}")

    if args.dry_run:
        log("Dry run — extracted data only.")
        crete_total = sum(v["pax_total"] for v in latest_data.values())
        log(f"  Crete total pax {MONTH_NAMES[latest_month]} {year}: {crete_total:,}")
        return 0

    conn = psycopg2.connect(**DB_PARAMS)
    try:
        ensure_schema(conn)

        # Persist all populated months for this year (helpful for context).
        for m in range(1, 13):
            month_data = parse_month_sheet(wb, m)
            for iata, vals in month_data.items():
                if vals["pax_total"] > 0:
                    upsert_airport_row(conn, year, m, iata, vals, url)

        if not args.force and already_in_ledger(conn, year, latest_month):
            log(f"  article {year}-{latest_month:02d} already published — skip")
            return 0

        history = fetch_year_history(conn, year)
        log(f"  YTD history loaded: {len(history)} months")

        prompt = build_prompt(year, latest_month, latest_data, history, url)
        log(f"  calling Claude (haiku) — prompt {len(prompt)} chars")
        t0 = time.time()
        raw = call_claude(prompt, model="haiku")
        article = parse_json_response(raw)
        log(f"  Claude returned in {time.time()-t0:.1f}s, confidence={article.get('confidence')}")

        if (article.get("confidence") or "").lower() == "low":
            log("  confidence=low — abort.")
            return 0

        guide_id = insert_draft(conn, article, year, latest_month, url)
        slug = derive_slug(article, year, latest_month)
        record_ledger(conn, year, latest_month, guide_id, slug)
        log(f"  inserted draft guide_id={guide_id} title='{article.get('title_en','')[:90]}'")

        crete_total = sum(v["pax_total"] for v in latest_data.values())
        send_telegram(
            f"🛬 HCAA Crete article #{guide_id}\n"
            f"{MONTH_NAMES[latest_month]} {year}: {crete_total:,} pax across HER+CHQ+JSH\n"
            f"{article.get('title_en','')[:120]}\n"
            f"https://crete.direct/en/news"
        )
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        log("Interrupted.")
        sys.exit(130)
