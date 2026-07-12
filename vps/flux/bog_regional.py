#!/usr/bin/env python3
"""Importe les indicateurs annuels Crete de la Banque de Grece.

Les fichiers officiels exposent les visites, nuitees, recettes, depenses et
duree de sejour par region. Seule la ligne agregee CRETE (EL43) est stockee.
Dependances : requests, openpyxl.
"""
import argparse
import io
import sys
from datetime import date

import requests
from openpyxl import load_workbook

try:  # execution directe VPS et import package dans pytest
    from .db import connect
except ImportError:
    from db import connect

SOURCES = {
    "receipts": ("https://www.bankofgreece.gr/RelatedDocuments/Receipts_regional.xlsx", "million_eur"),
    "visits": ("https://www.bankofgreece.gr/RelatedDocuments/Visits_regional.xlsx", "thousands"),
    "overnight_stays": ("https://www.bankofgreece.gr/RelatedDocuments/Overnight_stays_regional.xlsx", "thousands"),
    "expenditure_per_visit": ("https://www.bankofgreece.gr/RelatedDocuments/Expediture_per_visit_%20regional.xlsx", "eur"),
    "expenditure_per_night": ("https://www.bankofgreece.gr/RelatedDocuments/Expenditure_per_overnight_stay_regional.xlsx", "eur"),
    "average_stay": ("https://www.bankofgreece.gr/RelatedDocuments/Average_lenght_of%20_stay_regional.xlsx", "nights"),
}
UA = "crete.direct flux research/1.0 (official statistics importer)"


def extract_crete_annual(content):
    sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    crete_idx = next(i for i, row in enumerate(rows) if str(row[0] or "").strip().upper().startswith("CRETE (EL43)"))
    header = next(
        row for row in rows[:crete_idx]
        if str(row[0] or "").strip().lower().startswith("region")
        and sum(isinstance(value, (int, float)) and 2000 <= int(value) <= date.today().year for value in row) >= 3
    )
    values = rows[crete_idx]
    result = []
    started = False
    previous_year = None
    for col, value in enumerate(header):
        is_year = isinstance(value, (int, float)) and 2000 <= int(value) <= date.today().year
        if not started:
            if not is_year:
                continue
            started = True
        if not is_year or (previous_year is not None and int(value) <= previous_year):
            break  # fin du bloc annuel ; les colonnes suivantes sont trimestrielles
        metric_value = values[col]
        if isinstance(metric_value, (int, float)):
            result.append((int(value), float(metric_value)))
        previous_year = int(value)
    return result


UPSERT = """
insert into flux_economy_annual (year, region, metric, value, unit, source_url)
values (%s, 'crete', %s, %s, %s, %s)
on conflict (year, region, metric) do update set
  value = excluded.value, unit = excluded.unit, source_url = excluded.source_url, updated_at = now()
"""


def collect():
    session = requests.Session()
    session.headers["User-Agent"] = UA
    records = []
    for metric, (url, unit) in SOURCES.items():
        response = session.get(url, timeout=60)
        response.raise_for_status()
        for year, value in extract_crete_annual(response.content):
            records.append((year, metric, value, unit, url))
    return records


def run(dry_run=False):
    rows = collect()
    if not rows:
        raise RuntimeError("aucune serie Banque de Grece trouvee")
    if dry_run:
        for row in rows:
            print("|".join(map(str, row[:4])))
        return len(rows)
    conn = connect()
    conn.autocommit = True
    with conn, conn.cursor() as cur:
        cur.executemany(UPSERT, rows)
    conn.close()
    print(f"bog_regional: {len(rows)} lignes upsert")
    return len(rows)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    try:
        run(args.dry_run)
    except Exception as exc:
        print(f"bog_regional ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
